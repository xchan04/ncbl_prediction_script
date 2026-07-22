"""Load results + rankings from the league Google Sheet.

Three ways in, all handled by `League.load`:
  * a downloaded workbook (.xlsx) or per-tab .csv (or a folder of CSVs), or
  * a **shareable link** — a Google Sheets URL (exported to xlsx on the fly) or a
    direct http(s) link to an .xlsx/.csv — so no manual download is needed
    (see `sheet_source.py`).

Names are normalized to lowercase keys (the sheet has casing typos like
"deviousSprite"/"DeviousSprite" that must be merged); a display-name map keeps
the nicest casing for labels.
"""
from __future__ import annotations
import csv
import os
from collections import defaultdict, OrderedDict

from . import points as P
from . import sheet_source as SS


class League:
    """Everything derived from the sheet: per-player event lists, chronology,
    display names, and (optionally) the published ranking."""

    def __init__(self, cfg):
        self.cfg = cfg
        self.events = []                       # list of dicts (one row per result)
        self.by_player = defaultdict(list)     # lc name -> [(t_index, points)]
        self.display = {}                      # lc name -> nicest casing
        self.tournaments = []                  # chronological tournament names
        self.t_index = {}                      # tournament name -> 1-based index
        self.last_event = {}                   # lc name -> last t_index seen
        self.published_rank = {}               # lc name -> int (from rankings tab)
        self.published_points = {}             # lc name -> float
        self.roster = set()                    # lc names eligible to be ranked

    # ---------- loading ----------
    def load(self, path, data_sheet=None, rankings_sheet=None):
        """`path` may be a local .xlsx/.csv/folder, OR a shareable sheet URL
        (Google Sheets link or a direct http(s) link to an .xlsx/.csv)."""
        tmp = None
        src = path
        if SS.is_url(path):
            src = tmp = SS.fetch(path)          # download to a temp file, load, then remove
        try:
            ext = os.path.splitext(src)[1].lower()
            if os.path.isdir(src) or ext == ".csv":
                self._load_csv_dir(src)
            elif ext in (".xlsx", ".xlsm"):
                self._load_xlsx(src, data_sheet or self.cfg["data_entry_sheet"],
                                rankings_sheet or self.cfg["rankings_sheet"])
            else:
                raise ValueError(f"Unsupported input: {path} (use .xlsx, a .csv, a folder of CSVs, or a sheet URL)")
        finally:
            if tmp and os.path.exists(tmp):
                os.remove(tmp)
        self._finalize()
        return self

    def _load_xlsx(self, path, data_sheet, rankings_sheet):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        col = self.cfg["columns"]
        ws = wb[data_sheet]
        for r in range(self.cfg["data_entry_header_rows"] + 1, ws.max_row + 1):
            self._add_row(
                ws.cell(row=r, column=col["ref"]).value,
                ws.cell(row=r, column=col["tournament"]).value,
                ws.cell(row=r, column=col["cap"]).value,
                ws.cell(row=r, column=col["player"]).value,
                ws.cell(row=r, column=col["placement"]).value,
                ws.cell(row=r, column=col["gs_wins"]).value,
                ws.cell(row=r, column=col["points"]).value,
            )
        if rankings_sheet and rankings_sheet in wb.sheetnames:
            self._load_rankings_ws(wb[rankings_sheet])

    def _load_csv_dir(self, path):
        """`path` can be a single Data-Entry CSV, or a directory containing both
        the Data-Entry and Solo-Rankings CSVs (matched by filename keywords)."""
        col = self.cfg["columns"]
        files = [path]
        if os.path.isdir(path):
            files = sorted(os.path.join(path, f) for f in os.listdir(path) if f.lower().endswith(".csv"))
        for f in files:
            low = os.path.basename(f).lower()
            with open(f, newline="") as fh:
                rows = list(csv.reader(fh))
            if "data entry" in low or "data_entry" in low or len(files) == 1:
                for r in rows[self.cfg["data_entry_header_rows"]:]:
                    if len(r) < max(col.values()):
                        continue
                    self._add_row(r[col["ref"]-1], r[col["tournament"]-1], r[col["cap"]-1],
                                  r[col["player"]-1], r[col["placement"]-1],
                                  r[col["gs_wins"]-1], r[col["points"]-1])
            if "rankings" in low:
                self._load_rankings_rows(rows)

    def _add_row(self, ref, tour, cap, player, placement, gs, pts):
        if not player or pts in (None, "", " "):
            return
        try:
            pts = float(pts)
        except (TypeError, ValueError):
            return
        try:
            ref = float(ref)
        except (TypeError, ValueError):
            ref = len(self.events) + 1
        lc = str(player).strip().lower()
        self.display.setdefault(lc, str(player).strip())
        self.events.append({"ref": ref, "tournament": str(tour).strip() if tour else "",
                            "cap": cap, "player": lc, "placement": placement,
                             "gs": gs, "points": pts})

    def _load_rankings_ws(self, ws):
        rc = self.cfg["rankings_cols"]
        for r in range(self.cfg["rankings_header_rows"] + 1, ws.max_row + 1):
            self._add_rank(ws.cell(row=r, column=rc["rank"]).value,
                           ws.cell(row=r, column=rc["player"]).value,
                           ws.cell(row=r, column=rc["points"]).value)

    def _load_rankings_rows(self, rows):
        rc = self.cfg["rankings_cols"]
        for r in rows[self.cfg["rankings_header_rows"]:]:
            if len(r) < max(rc.values()):
                continue
            self._add_rank(r[rc["rank"]-1], r[rc["player"]-1], r[rc["points"]-1])

    def _add_rank(self, rank, name, pts):
        if not name or str(rank).strip() in ("", None):
            return
        try:
            rank = int(float(rank))
        except (TypeError, ValueError):
            return
        lc = str(name).strip().lower()
        self.display.setdefault(lc, str(name).strip())
        self.published_rank[lc] = rank
        try:
            self.published_points[lc] = float(pts)
        except (TypeError, ValueError):
            self.published_points[lc] = 0.0

    def _finalize(self):
        order = OrderedDict()
        for e in sorted(self.events, key=lambda e: e["ref"]):
            if e["tournament"] not in order:
                order[e["tournament"]] = e["ref"]
        self.tournaments = list(order.keys())
        self.t_index = {t: i + 1 for i, t in enumerate(self.tournaments)}
        for e in self.events:
            ti = self.t_index[e["tournament"]]
            self.by_player[e["player"]].append((ti, e["points"]))
            self.last_event[e["player"]] = max(self.last_event.get(e["player"], 0), ti)
        for p in self.by_player:
            self.by_player[p].sort()
        # roster = the universe of players eligible to be ranked.
        # With ranked_only, restrict to those in the rankings tab (registered players).
        if self.cfg.get("ranked_only", True) and self.published_rank:
            self.roster = {p for p in self.by_player if p in self.published_rank}
        else:
            self.roster = set(self.by_player.keys())

    # ---------- helpers ----------
    def name(self, lc):
        return self.display.get(lc, lc)

    def points_through(self, player, upto):
        evs = [pt for i, pt in self.by_player[player] if i <= upto]
        return P.season_score(evs, self.cfg)

    def score(self, player):
        return self.points_through(player, len(self.tournaments))

    def n_events(self, player):
        return len(self.by_player[player])
